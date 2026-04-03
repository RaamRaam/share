    """
TestReporter — captures step-level results per test case and writes an Excel report.

Excel output has three sheets:
  • Summary  — one row per test case   (TC#, scenario, total steps, pass, fail, result, duration, error)
  • Details  — one row per step        (TC#, scenario, step#, description, result, error)
  • Step Plan — steps listed up-front  (TC#, scenario, step#, planned step description)

Usage pattern (declarative):
──────────────────────────────
    def tc_001_open_app(page, reporter):
        reporter.start_test("TC-001", "Open application and verify home page")

        steps = [
            ("Navigate to APP_URL",          lambda: navigate(page, APP_URL)),
            ("Verify URL matches APP_URL",   lambda: check_url(page, APP_URL)),
            ("Verify page header visible",   lambda: check_element_visible(page, "h1")),
        ]

        return reporter.run_steps(steps)
"""

import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── colour palette ──────────────────────────────────────────────────────────
_HEADER_FILL   = PatternFill("solid", fgColor="2E4057")
_PASS_FILL     = PatternFill("solid", fgColor="C6EFCE")
_FAIL_FILL     = PatternFill("solid", fgColor="FFC7CE")
_ALT_FILL      = PatternFill("solid", fgColor="F2F2F2")
_HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
_PASS_FONT     = Font(bold=True, color="276221")
_FAIL_FONT     = Font(bold=True, color="9C0006")
_THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


class TestReporter:
    def __init__(self):
        self._results: list[dict] = []
        self._current: dict | None = None

    # ── lifecycle ────────────────────────────────────────────────────────────
    def start_test(self, tc_num: str, scenario: str) -> None:
        """Call at the beginning of each test-case function."""
        self._current = {
            "tc_num":   tc_num,
            "scenario": scenario,
            "steps":    [],
            "result":   "PASS",
            "error":    "",
            "start":    time.time(),
        }

    def end_test(self) -> None:
        """Call at the very end of a test-case function (pass or fail path)."""
        if self._current:
            self._current["duration"] = round(time.time() - self._current["start"], 2)
            self._results.append(self._current)
            self._current = None

    # ── step logging ─────────────────────────────────────────────────────────
    def log_step(self, description: str, result: str = "PASS", error: str = "") -> None:
        """Log a single step inside the current test case."""
        if not self._current:
            raise RuntimeError("Call start_test() before log_step()")
        step_num = len(self._current["steps"]) + 1
        self._current["steps"].append({
            "step_num":    step_num,
            "description": description,
            "result":      result,
            "error":       error,
        })
        if result == "FAIL":
            self._current["result"] = "FAIL"
            if not self._current["error"]:
                self._current["error"] = error

    def fail_test(self, description: str, error: str) -> None:
        """Convenience: log a FAIL step and mark the test case as failed."""
        self.log_step(description, "FAIL", error)

    # ── declarative step runner ───────────────────────────────────────────────
    def run_steps(self, steps: list[tuple]) -> bool:
        """
        Execute a declarative list of (description, callable) steps in order.

        • Each callable is called with no arguments (use lambdas to capture page etc.).
        • If a callable returns False explicitly → logged as FAIL.
        • Any exception → logged as FAIL with the exception message.
        • Calls end_test() automatically when done.
        • Returns True if every step passed, False otherwise.

        Example:
            steps = [
                ("Navigate to URL",        lambda: navigate(page, URL)),
                ("Check URL",              lambda: check_url(page, URL)),
                ("Verify header visible",  lambda: check_element_visible(page, "h1")),
            ]
            return reporter.run_steps(steps)
        """
        if not self._current:
            raise RuntimeError("Call start_test() before run_steps()")

        # Record the planned steps list so the Step Plan sheet can be written
        self._current["planned"] = [desc for desc, _ in steps]

        all_passed = True
        try:
            for description, action in steps:
                try:
                    result = action()
                    if result is False:
                        raise AssertionError("Action returned False")
                    self.log_step(description, "PASS")
                except Exception as exc:
                    self.log_step(description, "FAIL", str(exc))
                    all_passed = False
        finally:
            self.end_test()

        return all_passed

    # ── excel export ─────────────────────────────────────────────────────────
    def to_excel(self, path: str = "test_results.xlsx") -> str:
        wb = Workbook()

        self._write_summary_sheet(wb)
        self._write_details_sheet(wb)
        self._write_step_plan_sheet(wb)

        wb.save(path)
        print(f"\n✅  Report saved → {path}")
        return path

    # ── private helpers ───────────────────────────────────────────────────────
    def _write_summary_sheet(self, wb: Workbook) -> None:
        ws = wb.active
        ws.title = "Summary"

        headers = ["TC #", "Scenario", "Total Steps", "Passed", "Failed", "Result", "Duration (s)", "Error"]
        col_widths = [10, 45, 13, 10, 10, 10, 14, 50]

        self._write_header_row(ws, headers, col_widths)

        for row_idx, tc in enumerate(self._results, start=2):
            steps      = tc["steps"]
            passed     = sum(1 for s in steps if s["result"] == "PASS")
            failed     = len(steps) - passed
            is_pass    = tc["result"] == "PASS"
            row_fill   = _PASS_FILL if is_pass else _FAIL_FILL
            alt_fill   = _ALT_FILL if row_idx % 2 == 0 else PatternFill()

            values = [
                tc["tc_num"],
                tc["scenario"],
                len(steps),
                passed,
                failed,
                tc["result"],
                tc.get("duration", ""),
                tc["error"],
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                # colour the Result cell; other cells get alternating row shading
                if col_idx == 6:
                    cell.fill = row_fill
                    cell.font = _PASS_FONT if is_pass else _FAIL_FONT
                else:
                    cell.fill = alt_fill

        ws.freeze_panes = "A2"

    def _write_details_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Details")

        headers = ["TC #", "Scenario", "Step #", "Step Description", "Result", "Error"]
        col_widths = [10, 40, 8, 60, 10, 50]

        self._write_header_row(ws, headers, col_widths)

        row_idx = 2
        for tc in self._results:
            for step in tc["steps"]:
                is_pass = step["result"] == "PASS"
                values = [
                    tc["tc_num"],
                    tc["scenario"],
                    step["step_num"],
                    step["description"],
                    step["result"],
                    step["error"],
                ]
                for col_idx, val in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = _THIN_BORDER
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                    if col_idx == 5:
                        cell.fill = _PASS_FILL if is_pass else _FAIL_FILL
                        cell.font = _PASS_FONT if is_pass else _FAIL_FONT
                    elif row_idx % 2 == 0:
                        cell.fill = _ALT_FILL
                row_idx += 1

        ws.freeze_panes = "A2"

    def _write_step_plan_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Step Plan")

        headers = ["TC #", "Scenario", "Step #", "Planned Step Description"]
        col_widths = [10, 40, 8, 70]

        self._write_header_row(ws, headers, col_widths)

        row_idx = 2
        for tc in self._results:
            planned = tc.get("planned") or [s["description"] for s in tc["steps"]]
            for step_num, desc in enumerate(planned, start=1):
                values = [tc["tc_num"], tc["scenario"], step_num, desc]
                for col_idx, val in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = _THIN_BORDER
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                    if row_idx % 2 == 0:
                        cell.fill = _ALT_FILL
                row_idx += 1

        ws.freeze_panes = "A2"

    @staticmethod
    def _write_header_row(ws, headers: list[str], col_widths: list[int]) -> None:
        for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.font      = _HEADER_FONT
            cell.fill      = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 20
